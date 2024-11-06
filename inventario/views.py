from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.shortcuts import render, get_object_or_404, redirect
from django.db import transaction
from django.http import JsonResponse
from django.utils import timezone
from django.db.models import Sum, F
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.template.loader import get_template
from django.db.models import Q
from openpyxl import Workbook
from xhtml2pdf import pisa
from .forms import LoginForm, ProductoForm, ClienteForm, Cliente, CategoriaForm, CategoriaProducto, EmpleadoForm, ProveedorForm
from .models import Empleado, Proveedor, Producto, Venta, DetalleVenta, Compra, DetalleCompra, Inventario, Cliente, CategoriaProducto
import json
import datetime

def crear_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('productos_listar')  # Redirigir a la lista de productos 
    else:
        form = ProductoForm()
    return render(request, 'inventario/crear_producto.html', {'form': form})

def listar_productos(request):
    productos = Producto.objects.all()  # Obtiene todos los productos
    return render(request, 'inventario/listar_productos.html', {'productos': productos})

def crear_cliente(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('clientes_listar')  # Redirigir a la lista de clientes 
        form = ClienteForm()
    return render(request, 'inventario/crear_cliente.html', {'form': form})

def listar_clientes(request): #Sirve para ver la lista de clientes que hay
    clientes = Cliente.objects.all()  #Obtiene todos los clientes de la base de datos
    return render(request, 'inventario/listar_clientes.html', {'clientes': clientes})  #Renderiza el template con los clientes

def crear_categoria(request):
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            form.save()  # Guarda la nueva categoría en la base de datos
            return redirect('categorias_listar')  # Redirige a la lista de categorías
    else:
        form = CategoriaForm()
    return render(request, 'inventario/crear_categoria.html', {'form': form})


from .models import CategoriaProducto

def listar_categorias(request):
    categorias = CategoriaProducto.objects.all()
    return render(request, 'inventario/listar_categorias.html', {'categorias': categorias})

def crear_empleado(request):
    if request.method == 'POST':
        form = EmpleadoForm(request.POST)
        if form.is_valid():
            form.save()  # Guarda el nuevo empleado en la base de datos
            return redirect('empleados_listar')  # Redirige a la lista de empleados después de guardar
    else:
        form = EmpleadoForm()
    return render(request, 'inventario/crear_empleado.html', {'form': form})

def listar_empleados(request):
    empleados = Empleado.objects.all()  # Obtiene todos los empleados
    return render(request, 'inventario/listar_empleados.html', {'empleados': empleados})

def crear_proveedor(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            form.save()  # Guarda el nuevo proveedor en la base de datos
            return redirect('proveedores_listar')  # Redirige a la lista de proveedores después de guardar
    else:
        form = ProveedorForm()
    return render(request, 'inventario/crear_proveedor.html', {'form': form})

def listar_proveedores(request):
    proveedores = Proveedor.objects.all()  # Obtiene todos los proveedores
    return render(request, 'inventario/listar_proveedores.html', {'proveedores': proveedores})

def inicio(request):
    return render(request, 'inventario/inicio.html')


def venta_view(request):
    productos = Producto.objects.all()
    clientes = Cliente.objects.all()
    empleados = Empleado.objects.all()  # Obtener todos los empleados

    return render(request, 'inventario/venta.html', {
        'productos': productos,
        'clientes': clientes,
        'empleados': empleados  
    })


def compra_view(request):
    productos = Producto.objects.all() # Obtiene todos los productos
    proveedores = Proveedor.objects.all() # Obtiene todos los proveedores
    empleados = Empleado.objects.all()  # Obtiene todos los empleados

    return render(request, 'inventario/compra.html', {
        'productos': productos,
        'proveedores': proveedores,
        'empleados': empleados
    })


@csrf_exempt  
@transaction.atomic
def completar_compra(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            proveedor_id = data.get('proveedor')
            empleado_id = data.get('empleado')  # Obtener el empleado desde los datos JSON
            productos_compra = data.get('productos', [])
            total_compra = data.get('total')

            # Validar el proveedor
            try:
                proveedor = Proveedor.objects.get(nit_proveedor=proveedor_id)
            except Proveedor.DoesNotExist:
                return JsonResponse({'error': 'Proveedor no válido'}, status=400)

            # Validar el empleado
            try:
                empleado = Empleado.objects.get(id_usuario=empleado_id)
            except Empleado.DoesNotExist:
                return JsonResponse({'error': 'Empleado no válido'}, status=400)

            # Crear la compra con el proveedor y el empleado
            compra = Compra.objects.create(
                nit_proveedor=proveedor,
                id_usuario=empleado,  
                fecha_compra=timezone.now(),
                total_compra=total_compra
            )

            # Procesar los productos y actualizar el inventario
            for producto_data in productos_compra:
                producto_nombre = producto_data['nombre']
                fecha_expiracion = producto_data['fecha_expiracion']  # Fecha de caducidad
                cantidad = int(producto_data['cantidad'])
                precio = float(producto_data['precio'])

                try:
                    producto = Producto.objects.get(nombre_producto=producto_nombre)
                except Producto.DoesNotExist:
                    return JsonResponse({'error': f'Producto con nombre {producto_nombre} no existe'}, status=400)

                # Detalle de la compra
                DetalleCompra.objects.create(
                    id_compra=compra,
                    id_producto=producto,
                    cantidad_producto=cantidad,
                    total_producto=precio * cantidad,
                    fecha_expiracion=fecha_expiracion  # Guardamos la fecha de expiración aquí
                )

                # Actualizar o crear un registro en el inventario
                inventario_item, created = Inventario.objects.get_or_create(
                    id_producto=producto,
                    fecha_expiracion=fecha_expiracion,  # Asocia la fecha de expiración
                    defaults={'cantidad_disponible': 0}
                )

                if created:
                    inventario_item.cantidad_disponible = cantidad
                else:
                    inventario_item.cantidad_disponible += cantidad

                inventario_item.save()

            return JsonResponse({'success': 'Compra completada y inventario actualizado'})

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)

@csrf_exempt  
@transaction.atomic
def completar_venta(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            cliente_id = data.get('cliente')
            empleado_id = data.get('empleado')  # Obtener el empleado que realiza la venta
            productos_venta = data.get('productos', [])
            total_venta = data.get('total')

            # Validar el cliente
            try:
                cliente = Cliente.objects.get(nit_cliente=cliente_id)
            except Cliente.DoesNotExist:
                return JsonResponse({'error': 'Cliente no válido'}, status=400)

            # Validar el empleado
            try:
                empleado = Empleado.objects.get(id_usuario=empleado_id)
            except Empleado.DoesNotExist:
                return JsonResponse({'error': 'Empleado no válido'}, status=400)

            # Crear la venta
            venta = Venta.objects.create(
                nit_cliente=cliente,
                id_usuario=empleado,
                fecha_venta=timezone.now(),
                total_venta=total_venta
            )

            # Procesar cada producto en la venta
            for producto_data in productos_venta:
                producto_nombre = producto_data['nombre']
                cantidad_vendida = int(producto_data['cantidad'])
                precio = float(producto_data['precio'])

                # Obtener el producto por nombre
                try:
                    producto = Producto.objects.get(nombre_producto=producto_nombre)
                except Producto.DoesNotExist:
                    return JsonResponse({'error': f'Producto con nombre {producto_nombre} no existe'}, status=400)

                # Crear el detalle de la venta
                DetalleVenta.objects.create(
                    id_venta=venta,
                    id_producto=producto,
                    cantidad_producto=cantidad_vendida,
                    total_producto=precio * cantidad_vendida
                )

                # Actualizar el inventario, empezando por el producto más próximo a vencer
                inventarios = Inventario.objects.filter(id_producto=producto, cantidad_disponible__gt=0).order_by(
                    'fecha_expiracion')

                for inventario_item in inventarios:
                    if cantidad_vendida <= 0:
                        break 

                    if inventario_item.cantidad_disponible >= cantidad_vendida:
                        # Si el inventario actual tiene suficiente cantidad para cubrir la venta
                        inventario_item.cantidad_disponible -= cantidad_vendida
                        inventario_item.save()
                        cantidad_vendida = 0
                    else:
                        # Si el inventario actual no es suficiente, reducir lo que se pueda y continuar con el siguiente
                        cantidad_vendida -= inventario_item.cantidad_disponible
                        inventario_item.cantidad_disponible = 0
                        inventario_item.save()

            return JsonResponse({'success': 'Venta completada y inventario actualizado'})

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)


def inventario_view(request):
    # Filtrar todos los productos del inventario
    inventario = Inventario.objects.all()

    # Eliminar productos con cantidad disponible igual a 0
    inventario.filter(cantidad_disponible=0).delete()

    # Refrescar el inventario después de eliminar los productos sin existencias
    inventario = Inventario.objects.all().order_by('id_producto', 'fecha_expiracion')

    return render(request, 'inventario/inventario.html', {'inventario': inventario})

@login_required
def listar_ventas(request):
    ventas = Venta.objects.all()  # Obtiene todas las ventas
    return render(request, 'inventario/listar_ventas.html', {'ventas': ventas})


@login_required
def detalle_venta(request, venta_id):
    venta = get_object_or_404(Venta, id_venta=venta_id)  # Obtiene la venta específica por su ID
    detalles = DetalleVenta.objects.filter(id_venta=venta)  # Obtiene todos los productos en esa venta

    return render(request, 'inventario/detalle_venta.html', {
        'venta': venta,
        'detalles': detalles
    })
@login_required
def listar_categorias(request):
    categorias = CategoriaProducto.objects.all()
    return render(request, 'inventario/listar_categorias.html', {'categorias': categorias})


@login_required
def editar_categoria(request, categoria_id):
    categoria = get_object_or_404(CategoriaProducto, id_categoria=categoria_id)

    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría actualizada correctamente.')
            return redirect('categorias_listar')
    else:
        form = CategoriaForm(instance=categoria)

    return render(request, 'inventario/editar_categoria.html', {'form': form})


@login_required
@csrf_exempt
def eliminar_categoria(request, categoria_id):
    if request.method == 'POST':
        categoria = get_object_or_404(CategoriaProducto, id_categoria=categoria_id)
        categoria.delete()
        return JsonResponse({'success': True})  # Devolver JSON para confirmar la eliminación
    return JsonResponse({'error': 'Método no permitido'}, status=405)


#EDITAR CLIENTE

@login_required
def editar_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, nit_cliente=cliente_id)

    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cliente actualizado correctamente.')
            return redirect('clientes_listar')
    else:
        form = ClienteForm(instance=cliente)

    return render(request, 'inventario/editar_cliente.html', {'form': form})


@login_required
@csrf_exempt
def eliminar_cliente(request, cliente_id):
    if request.method == 'POST':
        cliente = get_object_or_404(Cliente, nit_cliente=cliente_id)
        cliente.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'error': 'Método no permitido'}, status=405)


#LISTADO DE COMPRAS
@login_required
def listar_compras(request):
    compras = Compra.objects.all()  # Obtiene todas las compras
    return render(request, 'inventario/listar_compras.html', {'compras': compras})

@login_required
def detalle_compra(request, compra_id):
    compra = get_object_or_404(Compra, id_compra=compra_id)  # Obtiene la compra específica por su ID
    detalles = DetalleCompra.objects.filter(id_compra=compra)  # Obtiene todos los productos en esa compra

    return render(request, 'inventario/detalle_compra.html', {
        'compra': compra,
        'detalles': detalles
    })

#LISTADO DE EMPLEADOS
@login_required
def editar_empleado(request, empleado_id):
    empleado = get_object_or_404(Empleado, id_usuario=empleado_id)

    if request.method == 'POST':
        form = EmpleadoForm(request.POST, instance=empleado)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empleado actualizado correctamente.')
            return redirect('empleados_listar')
    else:
        form = EmpleadoForm(instance=empleado)

    return render(request, 'inventario/editar_empleado.html', {'form': form})


@login_required
@csrf_exempt
def eliminar_empleado(request, empleado_id):
    if request.method == 'POST':
        empleado = get_object_or_404(Empleado, id_usuario=empleado_id)
        empleado.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'error': 'Método no permitido'}, status=405)


#EDITAR PRODUCTOS
@login_required
def editar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id_producto=producto_id)

    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, 'Producto actualizado correctamente.')
            return redirect('productos_listar')
    else:
        form = ProductoForm(instance=producto)

    return render(request, 'inventario/editar_producto.html', {'form': form})


@login_required
@csrf_exempt
def eliminar_producto(request, producto_id):
    if request.method == 'POST':
        producto = get_object_or_404(Producto, id_producto=producto_id)
        producto.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'error': 'Método no permitido'}, status=405)

#EDITAR PROVEEDORES
@login_required
def editar_proveedor(request, proveedor_id):
    proveedor = get_object_or_404(Proveedor, nit_proveedor=proveedor_id)

    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Proveedor actualizado correctamente.')
            return redirect('proveedores_listar')
    else:
        form = ProveedorForm(instance=proveedor)

    return render(request, 'inventario/editar_proveedor.html', {'form': form})


@login_required
@csrf_exempt
def eliminar_proveedor(request, proveedor_id):
    if request.method == 'POST':
        proveedor = get_object_or_404(Proveedor, nit_proveedor=proveedor_id)
        proveedor.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'error': 'Método no permitido'}, status=405)

def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('inicio')  # Redirige a la página de inicio si el login es exitoso
            else:
                messages.error(request, 'Usuario o contraseña incorrectos.')
    else:
        form = LoginForm()

    return render(request, 'inventario/login.html', {'form': form})

@login_required
def inicio_view(request):
    user = request.user

    # Verifica si el usuario pertenece al grupo "Empleados"
    es_empleado = user.groups.filter(name="Empleados").exists()

    return render(request, 'inventario/inicio.html', {'es_empleado': es_empleado})

@login_required
def cerrar_sesion(request):
    logout(request)
    return redirect('login')  # Redirige a la página de login después de cerrar sesión

#GENERAR REPORTE DE VENTAS
def reporte_ventas(request):
    # Obtener los parámetros de los filtros
    producto_id = request.GET.get('producto')
    empleado_id = request.GET.get('empleado')
    cliente_nit = request.GET.get('cliente')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    formato = request.GET.get('formato')  # 'excel' o 'pdf'

    # Aplicar filtros en la consulta de ventas
    ventas = Venta.objects.all()

    # Filtro por producto específico en los detalles de venta
    if producto_id:
        ventas = ventas.filter(detalleventa__id_producto=producto_id)

    # Filtros adicionales
    if empleado_id:
        ventas = ventas.filter(id_usuario__id_usuario=empleado_id)
    if cliente_nit:
        ventas = ventas.filter(nit_cliente__nit_cliente=cliente_nit)
    if fecha_inicio:
        ventas = ventas.filter(fecha_venta__gte=fecha_inicio)
    if fecha_fin:
        ventas = ventas.filter(fecha_venta__lte=fecha_fin)

    # Exportar en el formato seleccionado
    if formato == 'excel':
        return exportar_reporte_excel(ventas, producto_id)
    elif formato == 'pdf':
        return exportar_reporte_pdf(ventas, producto_id)
    else:
        # Renderizar la página de reportes con los datos y filtros
        productos = Producto.objects.all()
        empleados = Empleado.objects.all()
        clientes = Cliente.objects.all()
        return render(request, 'inventario/reporte_ventas.html', {
            'ventas': ventas,
            'productos': productos,
            'empleados': empleados,
            'clientes': clientes,
        })


def exportar_reporte_excel(ventas, producto_id=None):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Reporte de Ventas'

    # Encabezados de columna
    headers = ['ID Venta', 'Fecha Venta', 'Cliente', 'Empleado', 'Producto', 'Cantidad', 'Total']
    worksheet.append(headers)

    # Variable para almacenar el total de todas las ventas
    total_ventas = 0

    # Filas de datos
    for venta in ventas:
        detalles = venta.detalleventa_set.all()

        # Si se aplica filtro de producto, filtrar los detalles de venta
        if producto_id:
            detalles = detalles.filter(id_producto__id_producto=producto_id)

        for detalle in detalles:
            row = [
                venta.id_venta,
                venta.fecha_venta,
                venta.nit_cliente.nombre_completo,
                venta.id_usuario.nombre_usuario,
                detalle.id_producto.nombre_producto,
                detalle.cantidad_producto,
                f"Q{detalle.total_producto:.2f}"  # Prefijo "Q" agregado
            ]
            worksheet.append(row)

            # Sumar el total de cada detalle al total general
            total_ventas += detalle.total_producto

    # Añadir fila de total de ventas al final
    worksheet.append([])  # Fila vacía para separación
    worksheet.append(['', '', '', '', '', 'Total de Ventas:', f"Q{total_ventas:.2f}"])  # Prefijo "Q" agregado al total

    # Respuesta de descarga de archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte_de_Ventas.xlsx'
    workbook.save(response)
    return response


def exportar_reporte_pdf(ventas, producto_id=None):
    template_path = 'inventario/reporte_ventas_pdf.html'

    # Filtrar detalles de venta según el producto si se aplica filtro
    ventas_filtradas = []
    total_ventas = 0  # Variable para almacenar el total de todas las ventas

    for venta in ventas:
        detalles = venta.detalleventa_set.all()

        if producto_id:
            detalles = detalles.filter(id_producto__id_producto=producto_id)

        # Solo agregamos la venta si tiene detalles que coincidan con el filtro
        if detalles.exists():
            ventas_filtradas.append({
                'venta': venta,
                'detalles': detalles
            })
            # Sumar los totales de cada detalle al total general
            for detalle in detalles:
                total_ventas += detalle.total_producto

    # Pasamos total_ventas con el prefijo "Q" al contexto para usarlo en el template
    context = {
        'ventas_filtradas': ventas_filtradas,
        'total_ventas': f"Q{total_ventas:.2f}"  # Prefijo "Q" agregado al total
    }

    # Renderizar la plantilla en PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=Reporte_de_Ventas.pdf'
    template = get_template(template_path)
    html = template.render(context)
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('Hubo un error al generar el PDF', status=500)
    return response

#REPORTE DE COMPRAS

def reporte_compras(request):
    # Obtener los parámetros de los filtros
    producto_id = request.GET.get('producto')
    empleado_id = request.GET.get('empleado')
    proveedor_id = request.GET.get('proveedor')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    formato = request.GET.get('formato')  # 'excel' o 'pdf'

    # Aplicar filtros en la consulta de compras
    compras = Compra.objects.all()

    # Filtro por producto específico en los detalles de compra
    if producto_id:
        compras = compras.filter(detallecompra__id_producto=producto_id)

    # Filtros adicionales
    if empleado_id:
        compras = compras.filter(id_usuario__id_usuario=empleado_id)
    if proveedor_id:
        compras = compras.filter(nit_proveedor__nit_proveedor=proveedor_id)
    if fecha_inicio:
        compras = compras.filter(fecha_compra__gte=fecha_inicio)
    if fecha_fin:
        compras = compras.filter(fecha_compra__lte=fecha_fin)

    # Exportar en el formato seleccionado
    if formato == 'excel':
        return exportar_reporte_compras_excel(compras, producto_id)
    elif formato == 'pdf':
        return exportar_reporte_compras_pdf(compras, producto_id)
    else:
        # Renderizar la página de reportes con los datos y filtros
        productos = Producto.objects.all()
        empleados = Empleado.objects.all()
        proveedores = Proveedor.objects.all()
        return render(request, 'inventario/reporte_compras.html', {
            'compras': compras,
            'productos': productos,
            'empleados': empleados,
            'proveedores': proveedores,
        })


def exportar_reporte_compras_excel(compras, producto_id=None):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Reporte de Compras'

    # Encabezados de columna
    headers = ['ID Compra', 'Fecha Compra', 'Proveedor', 'Empleado', 'Producto', 'Cantidad', 'Total']
    worksheet.append(headers)

    # Variable para almacenar el total de todas las compras
    total_compras = 0

    # Filas de datos
    for compra in compras:
        detalles = compra.detallecompra_set.all()

        # Si se aplica filtro de producto, filtrar los detalles de compra
        if producto_id:
            detalles = detalles.filter(id_producto__id_producto=producto_id)

        for detalle in detalles:
            row = [
                compra.id_compra,
                compra.fecha_compra,
                compra.nit_proveedor.nombre_proveedor,
                compra.id_usuario.nombre_usuario,
                detalle.id_producto.nombre_producto,
                detalle.cantidad_producto,
                f"Q{detalle.total_producto:.2f}"  # Prefijo "Q" agregado
            ]
            worksheet.append(row)

            # Sumar el total de cada detalle al total general
            total_compras += detalle.total_producto

    # Añadir fila de total de compras al final
    worksheet.append([])  # Fila vacía para separación
    worksheet.append(['', '', '', '', '', 'Total de Compras:', f"Q{total_compras:.2f}"])  # Prefijo "Q" agregado al total

    # Respuesta de descarga de archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte_de_Compras.xlsx'
    workbook.save(response)
    return response


def exportar_reporte_compras_pdf(compras, producto_id=None):
    template_path = 'inventario/reporte_compras_pdf.html'

    # Filtrar detalles de compra según el producto si se aplica filtro
    compras_filtradas = []
    total_compras = 0  # Variable para almacenar el total de todas las compras

    for compra in compras:
        detalles = compra.detallecompra_set.all()

        if producto_id:
            detalles = detalles.filter(id_producto__id_producto=producto_id)

        # Solo agregamos la compra si tiene detalles que coincidan con el filtro
        if detalles.exists():
            compras_filtradas.append({
                'compra': compra,
                'detalles': detalles
            })
            # Sumar los totales de cada detalle al total general
            for detalle in detalles:
                total_compras += detalle.total_producto

    # Pasamos total_compras con el prefijo "Q" al contexto para usarlo en el template
    context = {
        'compras_filtradas': compras_filtradas,
        'total_compras': f"Q{total_compras:.2f}"  # Prefijo "Q" agregado al total
    }

    # Renderizar la plantilla en PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=Reporte_de_Compras.pdf'
    template = get_template(template_path)
    html = template.render(context)
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('Hubo un error al generar el PDF', status=500)
    return response

def reporte_inventario(request, formato):
    # Obtener los datos del inventario
    inventario = Inventario.objects.all().order_by('id_producto')

    if formato == 'excel':
        # Generar reporte en formato Excel
        return generar_reporte_excel(inventario)
    elif formato == 'pdf':
        # Generar reporte en formato PDF
        return generar_reporte_pdf(inventario)
    else:
        return HttpResponse("Formato de reporte no válido", status=400)

def generar_reporte_excel(inventario):
    # Crear el archivo Excel
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Reporte de Inventario'

    # Encabezados de columna
    headers = ['ID Producto', 'Nombre Producto', 'Fecha Expiración', 'Cantidad Disponible']
    worksheet.append(headers)

    # Agregar filas de datos
    for item in inventario:
        row = [
            item.id_producto.id_producto,
            item.id_producto.nombre_producto,
            item.fecha_expiracion,
            item.cantidad_disponible,
        ]
        worksheet.append(row)

    # Respuesta de descarga de archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte_de_Inventario.xlsx'
    workbook.save(response)
    return response

def generar_reporte_pdf(inventario):
    # Configurar la plantilla PDF
    template_path = 'inventario/reporte_inventario_pdf.html'
    context = {'inventario': inventario}

    # Renderizar la plantilla en PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=Reporte_de_Inventario.pdf'
    template = get_template(template_path)
    html = template.render(context)
    pisa_status = pisa.CreatePDF(html, dest=response)

    # Verificar si hubo algún error al generar el PDF
    if pisa_status.err:
        return HttpResponse('Hubo un error al generar el PDF', status=500)
    return response

def compra_rapida(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            cantidad_minima = int(data.get('cantidad_minima'))
            cantidad_a_comprar = int(data.get('cantidad_comprar'))
            fecha_vencimiento = data.get('fecha_vencimiento')

            # Obtener productos con cantidad en inventario menor o igual a la cantidad mínima
            productos_faltantes = Producto.objects.annotate(
                total_disponible=Sum('inventario__cantidad_disponible')
            ).filter(
                Q(total_disponible__lte=cantidad_minima) | Q(total_disponible__isnull=True)
            )

            productos_compra = []
            total_compra = 0

            # Procesar cada producto faltante
            for producto in productos_faltantes:
                total_a_comprar = cantidad_a_comprar
                precio = float(producto.precio_compra)
                total_producto = total_a_comprar * precio

                productos_compra.append({
                    'id_producto': producto.id_producto,
                    'nombre_producto': producto.nombre_producto,
                    'cantidad': total_a_comprar,
                    'precio': precio,
                    'total': total_producto,
                    'fecha_vencimiento': fecha_vencimiento
                })

                total_compra += total_producto

            return JsonResponse({
                'productos': productos_compra,
                'total_compra': total_compra
            })

        except Exception as e:
            print("Error en compra_rapida:", e)
            return JsonResponse({'error': str(e)}, status=500)

    # Renderizar la página de compra rápida
    proveedores = Proveedor.objects.all()
    empleados = Empleado.objects.all()
    return render(request, 'inventario/compra_rapida.html', {
        'proveedores': proveedores,
        'empleados': empleados
    })

@transaction.atomic
def completar_compra_rapida(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            proveedor_id = data.get('proveedor')
            empleado_id = data.get('empleado')
            productos = data.get('productos', [])

            # Validación de proveedor
            try:
                proveedor = Proveedor.objects.get(nit_proveedor=proveedor_id)
            except Proveedor.DoesNotExist:
                return JsonResponse({'error': 'Proveedor no válido'}, status=400)

            # Validación de empleado
            try:
                empleado = Empleado.objects.get(id_usuario=empleado_id)
            except Empleado.DoesNotExist:
                return JsonResponse({'error': 'Empleado no válido'}, status=400)

            # Verificar si la lista de productos está vacía
            if not productos:
                return JsonResponse({'error': 'No se enviaron productos para la compra'}, status=400)

            total_compra = 0

            # Crear la compra
            compra = Compra.objects.create(
                nit_proveedor=proveedor,
                id_usuario=empleado,
                fecha_compra=timezone.now(),
                total_compra=0  # Se actualizará luego con el total real
            )

            # Procesar cada producto
            for item in productos:
                producto_nombre = item.get('nombre_producto')
                cantidad = item.get('cantidad')
                precio = item.get('precio')
                fecha_vencimiento = item.get('fecha_vencimiento')

                # Validación de campos necesarios
                if not producto_nombre or not cantidad or not precio:
                    return JsonResponse({'error': 'Faltan datos en uno de los productos'}, status=400)

                # Obtener o crear el producto
                try:
                    producto = Producto.objects.get(nombre_producto=producto_nombre)
                except Producto.DoesNotExist:
                    return JsonResponse({'error': f'Producto {producto_nombre} no encontrado'}, status=400)

                # Calcular el total por producto y sumarlo al total general
                total_producto = float(precio) * int(cantidad)
                total_compra += total_producto

                # Crear el detalle de la compra
                DetalleCompra.objects.create(
                    id_compra=compra,
                    id_producto=producto,
                    cantidad_producto=cantidad,
                    total_producto=total_producto,
                    fecha_expiracion=fecha_vencimiento
                )

                # Actualizar o crear el inventario para el producto
                inventario, created = Inventario.objects.get_or_create(
                    id_producto=producto,
                    fecha_expiracion=fecha_vencimiento,
                    defaults={'cantidad_disponible': 0}
                )
                inventario.cantidad_disponible += int(cantidad)
                inventario.save()

            # Actualizar el total de la compra
            compra.total_compra = total_compra
            compra.save()

            return JsonResponse({'success': 'Compra rápida completada y productos añadidos al inventario'})

        except Exception as e:
            print("Error en completar_compra_rapida:", e)
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)